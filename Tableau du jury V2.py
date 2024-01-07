import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Demander à l'utilisateur de choisir l'année
annee = input("Pour quelle année souhaitez-vous créer le tableau de jury? Entrez 1 pour la première année ou 2 pour la deuxième année: ")

# Création d'un nouveau classeur Excel
new_workbook = Workbook()
new_sheet = new_workbook.active
if annee == "1":
    new_sheet.title = "Tableau jury I1"
elif annee == "2":
    new_sheet.title = "Tableau jury I2"

# Demander l'identifiant de l'élève et le chemin du fichier
identifiant_eleve = input("Veuillez entrer l'identifiant de l'élève : ")
chemin_fichier_moyennes = input("Veuillez entrer le chemin complet du fichier contenant les moyennes : ")

# Ouvrir le fichier de base de données
wb_moyennes = load_workbook(filename=chemin_fichier_moyennes)
feuille_info_etudiant = wb_moyennes.active  # Supposons que les infos sont dans la première feuille

# Configuration des styles des cellules
def configurer_styles(cell, row, col):
    # Style par défaut
    font = Font(name='Calibri', size=11, color="000000")
    fill = PatternFill(fill_type='solid', start_color='FFFFFF')
    alignment = Alignment(horizontal='left')
    border = Border()  # Aucune bordure par défaut

    # Bordures et style gras uniquement pour les lignes spécifiées
    if 10 <= row <= (36 if annee == "1" else 32) and 1 <= col <= 5:
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # Mettre en gras les lignes 1 et 10
    if row in [1, 10]:
        font = Font(name='Calibri', size=11, bold=True, color="000000")

    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

# Appliquer les styles aux cellules
for row in range(1, 100):
    for col in range(1, 25):  # Colonnes A à E
        cell_coordinate = f"{chr(64 + col)}{row}"
        cell = new_sheet[cell_coordinate]
        configurer_styles(cell, row, col)

# Ouvrir le fichier de base de données
wb_moyennes = load_workbook(filename=chemin_fichier_moyennes)

# Fonction pour colorer les cellules
def colorer_cellule(cell, moyenne):
    if moyenne < 10:
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rouge
    else:
        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Vert

# Parcourir chaque feuille pour trouver les moyennes de l'élève
moyennes_par_module = {}
rattrapages = []
modules_valides = []
somme_moyennes = 0
nombre_moyennes = 0
for nom_feuille in wb_moyennes.sheetnames:
    sheet = wb_moyennes[nom_feuille]
    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value).strip() == str(identifiant_eleve).strip():
            nom_module = row[4].value
            moyenne = row[3].value
            if nom_module and moyenne:
                moyennes_par_module[nom_module] = moyenne
                somme_moyennes += moyenne
                nombre_moyennes += 1
                if moyenne < 10:
                    rattrapages.append(nom_module)
                else:
                    modules_valides.append(nom_module)
            break

# Fermeture du workbook de base de données
wb_moyennes.close()

# Configuration du contenu en fonction de l'année
if annee == "1":
            
    # Contenu pour la première année
    content = {
    "A1": "Modules et sous modules I1",
    "A3": "Informations Étudiant",
    "A4": "Identifiant :",
    "A5": "Nom :",
    "A6": "Prénom :",
    "A7": "Classe :",
    "A10": "Code Modules",
    "A11": "CT.1208",
    "A13": "CS.2307",
    "A17": "CH.2307",
    "A22": "CH.2309",
    "A26": "CS.1106",
    "A28": "CS.1207",
    "A30": "CS.?",
    "A36": "AE.1101",
    "A40": "Rattrapages",

    "B10": "Modules",
    "B11": "SIGNAUX & SYSTEMES",
    "B13": "MATHEMATIQUES POUR L'INGENIEUR",
    "B17": "HUMANITES",
    "B22": "LANGUES VIVANTES",
    "B26": "ELECTROMAGNETISME I",
    "B28": "ELECTROMAGNETISME II",
    "B30": "PROJET BE",
    "B36": "APP Signal",
    "B40": "Modules validés",

    "C10": "Code sous modules",
    "C14": "CS.1108",
    "C15": "CS.1107",
    "C18": "CH.2307",
    "C19": "CH.1204",
    "C20": "CH.1106",
    "C23": "CH.1109",
    "C24": "LV2 CII",
    "C31": "CH.1109",
    "C32": "CH.1105",
    "C33": "CH.1111",
    "C34": "CH.1105",

    "D10": "Sous modules",
    "D14": "Mathématiques et informatique",
    "D15": "Mathématiques 1",
    "D18": "Communication",
    "D19": "Français",
    "D20": "Société environnement et entreprise ",
    "D23": "Anglais",
    "D24": "Espagnol",
    "D31": "BE Anglais",
    "D32": "BE Eco",
    "D33": "BE Data",
    "D34": "BE Robotique",

    "E10": "Moyennes",
    
    "F40": "Moyenne générale",
}   
    
elif annee == "2":
            
    # Contenu pour la deuxième année
    content = {
    "A1": "Modules et sous modules I2",
    "A3": "Informations Étudiant",
    "A4": "Identifiant :",
    "A5": "Nom :",
    "A6": "Prénom :",
    "A7": "Classe :",
    "A10": "Code Modules",
    "A11": "CT.2306",
    "A13": "CS.2307",
    "A17": "CH.2307",
    "A23": "CH.2309",
    "A28": "CS.2306",
    "A30": "CS.2309",
    "A32": "PTS",
    "A36": "Rattrapages",

    "B10": "Modules",
    "B11": "SIGNAUX & SYSTEMES II",
    "B13": "MATHEMATIQUES POUR L'INGENIEUR",
    "B17": "HUMANITES",
    "B23": "LANGUES VIVANTES & INTERCULTURALITES III",
    "B28": "VIBRATIONS & ONDES",
    "B30": "MECANIQUE QUANTIQUE",
    "B32": "PROJET TECHNIQUE SPECIALISE",
    "B36": "Modules validés",

    "C10": "Code sous modules",
    "C14": "CS.2307",
    "C15": "CS.2307",
    "C18": "CH.2307",
    "C19": "CH.2307",
    "C20": "CH.2307",
    "C21": "CH.2307",
    "C24": "CH.2309",
    "C25": "LV2 CII",
    "C26": "CH.2309",

    "D10": "Sous modules",
    "D14": "Méthodes Numériques",
    "D15": "Algèbre Linéaire",
    "D18": "Communication",
    "D19": "Français",
    "D20": "Fonctions des entr.",
    "D21": "Intro.Num.Respons.",
    "D24": "Anglais",
    "D25": "Espagnol",
    "D26": "Intercularité",

    "E10": "Moyennes",

    "F36": "Moyenne générale",
}
else:
    raise ValueError("Entrée invalide. Veuillez entrer 1 ou 2.")


for cell_coord, value in content.items():
    new_sheet[cell_coord].value = value

# Ajustement automatique de la largeur des colonnes jusqu'à la colonne F
for col_num in range(1, 7):  # Colonnes A à F
    max_length = 0
    col_letter = get_column_letter(col_num)
    for cell in new_sheet[col_letter]:
        try:
            cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)
        except:
            pass
    adjusted_width = (max_length + 2)
    new_sheet.column_dimensions[col_letter].width = adjusted_width

# Ajout du contenu spécifique (selon l'année choisie)
for cell_coord, value in content.items():
    new_sheet[cell_coord].value = value

# Intégrer les moyennes dans le nouveau classeur et colorer les cellules
ligne_rattrapages = 41 if annee == "1" else 37
ligne_modules_valides = ligne_rattrapages
for module, moyenne in moyennes_par_module.items():
    trouve = False
    for row in new_sheet.iter_rows(min_row=11, max_row=new_sheet.max_row, min_col=1, max_col=5):
        if row[0].value == module:
            cell_moyenne = row[4]
            cell_moyenne.value = moyenne
            colorer_cellule(cell_moyenne, moyenne)
            trouve = True
            break
    if not trouve:
        print(f"Moyenne pour le module '{module}' non trouvée dans le nouveau classeur.")


# Ajout des modules dans "Rattrapages" ou "Modules Validés"
for module in rattrapages:
    new_sheet[f"A{ligne_rattrapages}"].value = module
    ligne_rattrapages += 1
for module in modules_valides:
    new_sheet[f"B{ligne_modules_valides}"].value = module
    ligne_modules_valides += 1

def mettre_en_gras_ligne_complete(sheet, ligne):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=ligne, column=col)
        cell.font = Font(bold=True)

# Choix de la ligne à mettre en gras en fonction de l'année
ligne_a_mettre_en_gras = 40 if annee == "1" else 36

# Mettre en gras toute la ligne
mettre_en_gras_ligne_complete(new_sheet, ligne_a_mettre_en_gras)

def personnaliser_moyenne_generale(sheet, ligne, colonne):
    cell = sheet.cell(row=ligne, column=colonne)
    cell.font = Font(bold=True, color="0000FF")  # Texte en gras et en bleu
    cell.alignment = Alignment(horizontal="center", vertical="center")  # Texte centré

# Calcul et affichage de la moyenne générale
if nombre_moyennes > 0:
    moyenne_generale = somme_moyennes / nombre_moyennes
    ligne_moyenne_generale = 41 if annee == '1' else 37
    colonne_moyenne_generale = 6  # Colonne F
    new_sheet[f"F{ligne_moyenne_generale}"].value = moyenne_generale

    # Personnaliser l'apparence de la cellule de la moyenne générale
    personnaliser_moyenne_generale(new_sheet, ligne_moyenne_generale, colonne_moyenne_generale)


# Fonction pour trouver les informations de l'étudiant
def trouver_informations_etudiant(chemin_fichier, identifiant_eleve):
    wb = load_workbook(filename=chemin_fichier)
    feuille = wb.worksheets[0]  # Lecture de la première feuille

    for row in feuille.iter_rows(min_row=2):  # Commencer à lire à partir de la deuxième ligne
        identifiant = str(row[0].value)  # Convertir l'identifiant en chaîne de caractères
        if identifiant.strip() == str(identifiant_eleve).strip():
            nom = row[1].value
            prenom = row[2].value
            classe = row[3].value if len(row) > 3 else None
            return {'identifiant': identifiant, 'nom': nom, 'prenom': prenom, 'classe': classe}

    return None


infos_etudiant = trouver_informations_etudiant(chemin_fichier_moyennes, identifiant_eleve)
if infos_etudiant:
    new_sheet["A4"] = f"Identifiant : {infos_etudiant['identifiant']}"
    new_sheet["A5"] = f"Nom : {infos_etudiant['nom']}"
    new_sheet["A6"] = f"Prénom : {infos_etudiant['prenom']}"
    new_sheet["A7"] = f"Classe : {infos_etudiant['classe']}"
    print("Informations écrites dans le classeur Excel.")
else:
    print("Étudiant non trouvé.")



# Demande à l'utilisateur de saisir le nom du fichier
file_name = input("Veuillez entrer le nom du fichier à enregistrer (sans l'extension) : ") + '.xlsx'

# Détermination du chemin du bureau de l'utilisateur
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')

# Chemin complet pour enregistrer le fichier
new_file_path = os.path.join(desktop_path, file_name)

# Enregistrement du fichier sur le bureau
new_workbook.save(filename=new_file_path)

print(f"Le fichier a été enregistré sur le bureau sous : {new_file_path}")